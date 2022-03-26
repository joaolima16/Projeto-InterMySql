from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import json
wb = load_workbook('tabelaCarros.xlsx')
wb = wb.active

# Pergunta quantidade de colunas 
qtdCol = int(input("Insira a quantidade de colunas da tabela: "))
qtdLinhas = int(input("Insira a quantidade de linhas da tabela: "))

dados = {}

for col in range(0,qtdCol):
    letra = 65 + col
    letra = chr(letra)
    celula = f"{letra}1"
    informacoesCelula = []
    for lin in range(2,qtdLinhas+1):    
        valorCol = f"{letra}{lin}"
        informacoesCelula.append(wb[valorCol].value)
    dados[wb[celula].value] = informacoesCelula


dadosTratados = []
def formataJson():
    i = 0
    for linhas in range(1,qtdLinhas):
        arrayInfosTemp = []
        arrayIndex = []
        objetoLinha = {}
        for colunas in range(0,qtdCol):
            letra = 65 + colunas
            letra = chr(letra)
            celulaCampo = f"{letra}1"
            arrayIndex.append(wb[celulaCampo].value)
            arrayInfosTemp.append(dados[wb[celulaCampo].value][i])
        for infos in range(0,len(arrayIndex)):
            objetoLinha[arrayIndex[infos]] = arrayInfosTemp[infos]
        dadosTratados.append(objetoLinha)
        i = i + 1
        
formataJson()
arquivo = open("InfosExcel.json",'w')
json.dump(dadosTratados,arquivo,indent=4)